import os
import uuid
import json
from datetime import datetime, timedelta
from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

app = Flask(__name__)
app.secret_key = 'gstr3b-consolidator-secret-key'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'output'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB
ALLOWED_EXTENSIONS = {'json'}

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

def cleanup_old_files():
    now = datetime.now()
    for folder in [app.config['UPLOAD_FOLDER'], app.config['OUTPUT_FOLDER']]:
        for fname in os.listdir(folder):
            fpath = os.path.join(folder, fname)
            if os.path.isfile(fpath):
                mtime = datetime.fromtimestamp(os.path.getmtime(fpath))
                if now - mtime > timedelta(hours=1):
                    os.remove(fpath)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_nested(data, *keys, default=0):
    for key in keys:
        if isinstance(data, dict):
            data = data.get(key, {})
        else:
            return default
    return data if data is not None else default

def sum_array_field(data, array_key, field):
    arr = data.get(array_key, [])
    return sum(item.get(field, 0) for item in arr if isinstance(item, dict))

def parse_gstr3b_files(file_paths):
    months_data = {}
    gstin = None
    legal_name = ""
    months_list = []
    errors = []

    for fpath in file_paths:
        try:
            with open(fpath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            ret_period = data.get('ret_period')
            if not ret_period:
                errors.append(f"File {os.path.basename(fpath)} missing 'ret_period'")
                continue
            months_list.append(ret_period)
            if gstin is None:
                gstin = data.get('gstin', '')
            else:
                if data.get('gstin') != gstin:
                    errors.append(f"GSTIN mismatch in {os.path.basename(fpath)}")
            months_data[ret_period] = data
        except Exception as e:
            errors.append(f"Error reading {os.path.basename(fpath)}: {str(e)}")

    if not months_data:
        return None, "No valid JSON files found. " + "; ".join(errors)

    months_list.sort()
    first = months_list[0]
    last = months_list[-1]
    fy_start = first[:2] + first[2:4]
    fy_end = last[:2] + last[2:4]
    financial_year = f"{fy_start} - {fy_end}"
    no_of_months = len(months_list)

    # ------------------------------------------------------------------
    # Define all data rows (including totals) with extractors
    # Row numbers must match those used in template_rows (see create_excel_file)
    # ------------------------------------------------------------------
    rows = []

    def add_row(row_num, label, *path, default=0):
        rows.append({
            'row': row_num,
            'label': label,
            'extractor': lambda d, p=path: get_nested(d, *p, default=default)
        })

    # 3.1
    add_row(5, "(i) Total taxable value", "sup_details", "osup_det", "txval")
    add_row(6, "(ii) Integrated tax", "sup_details", "osup_det", "iamt")
    add_row(7, "(iii) Central tax", "sup_details", "osup_det", "camt")
    add_row(8, "(iv) State/UT tax", "sup_details", "osup_det", "samt")
    add_row(9, "(v) Cess", "sup_details", "osup_det", "csamt")
    add_row(11, "(i) Total taxable value", "sup_details", "osup_zero", "txval")
    add_row(12, "(ii) Integrated tax", "sup_details", "osup_zero", "iamt")
    add_row(13, "(v) Cess", "sup_details", "osup_zero", "csamt")
    add_row(15, "(i) Total taxable value", "sup_details", "osup_nil_exmp", "txval")
    add_row(17, "(i) Total taxable value", "sup_details", "isup_rev", "txval")
    add_row(18, "(ii) Integrated tax", "sup_details", "isup_rev", "iamt")
    add_row(19, "(iii) Central tax", "sup_details", "isup_rev", "camt")
    add_row(20, "(iv) State/UT tax", "sup_details", "isup_rev", "samt")
    add_row(21, "(v) Cess", "sup_details", "isup_rev", "csamt")
    add_row(23, "(i) Total taxable value", "sup_details", "osup_nongst", "txval")

    # 3.1.1
    add_row(28, "Total taxable value", "eco_dtls", "eco_sup", "txval")
    add_row(29, "Integrated tax", "eco_dtls", "eco_sup", "iamt")
    add_row(30, "Central tax", "eco_dtls", "eco_sup", "camt")
    add_row(31, "State/UT tax", "eco_dtls", "eco_sup", "samt")
    add_row(32, "Cess", "eco_dtls", "eco_sup", "csamt")
    add_row(34, "Total taxable value", "eco_dtls", "eco_reg_sup", "txval")
    # (ii) tax components not present; set zero
    rows.append({'row': 35, 'label': "Integrated tax", 'extractor': lambda d: 0})
    rows.append({'row': 36, 'label': "Central tax", 'extractor': lambda d: 0})
    rows.append({'row': 37, 'label': "State/UT tax", 'extractor': lambda d: 0})
    rows.append({'row': 38, 'label': "Cess", 'extractor': lambda d: 0})

    # 3.2 Inter State Supplies
    rows.append({'row': 43, 'label': "Supplies made to Unregistered Persons (Taxable value)",
                 'extractor': lambda d: sum_array_field(d.get('inter_sup', {}), 'unreg_details', 'txval')})
    rows.append({'row': 44, 'label': "Supplies made to Unregistered Persons (Integrated tax)",
                 'extractor': lambda d: sum_array_field(d.get('inter_sup', {}), 'unreg_details', 'iamt')})
    rows.append({'row': 46, 'label': "Supplies made to Composition Taxable Persons (Taxable value)",
                 'extractor': lambda d: sum_array_field(d.get('inter_sup', {}), 'comp_details', 'txval')})
    rows.append({'row': 47, 'label': "Supplies made to Composition Taxable Persons (Integrated tax)",
                 'extractor': lambda d: sum_array_field(d.get('inter_sup', {}), 'comp_details', 'iamt')})
    rows.append({'row': 49, 'label': "Supplies made to UIN holders (Taxable value)",
                 'extractor': lambda d: sum_array_field(d.get('inter_sup', {}), 'uin_details', 'txval')})
    rows.append({'row': 50, 'label': "Supplies made to UIN holders (Integrated tax)",
                 'extractor': lambda d: sum_array_field(d.get('inter_sup', {}), 'uin_details', 'iamt')})

    # 4. ITC Available
    def itc_avl_sum(ty, field):
        def extractor(d):
            avl = d.get('itc_elg', {}).get('itc_avl', [])
            return sum(item.get(field, 0) for item in avl if item.get('ty') == ty)
        return extractor

    rows.append({'row': 56, 'label': "Import of goods (Integrated tax)", 'extractor': itc_avl_sum('IMPG', 'iamt')})
    rows.append({'row': 57, 'label': "Import of goods (Cess)", 'extractor': itc_avl_sum('IMPG', 'csamt')})
    rows.append({'row': 59, 'label': "Import of services (Integrated tax)", 'extractor': itc_avl_sum('IMPS', 'iamt')})
    rows.append({'row': 60, 'label': "Import of services (Cess)", 'extractor': itc_avl_sum('IMPS', 'csamt')})
    rows.append({'row': 62, 'label': "Inward supplies liable to reverse charge (Integrated tax)", 'extractor': itc_avl_sum('ISRC', 'iamt')})
    rows.append({'row': 63, 'label': "Inward supplies liable to reverse charge (Central tax)", 'extractor': itc_avl_sum('ISRC', 'camt')})
    rows.append({'row': 64, 'label': "Inward supplies liable to reverse charge (State/UT tax)", 'extractor': itc_avl_sum('ISRC', 'samt')})
    rows.append({'row': 65, 'label': "Inward supplies liable to reverse charge (Cess)", 'extractor': itc_avl_sum('ISRC', 'csamt')})
    rows.append({'row': 67, 'label': "Inward supplies from ISD (Integrated tax)", 'extractor': itc_avl_sum('ISD', 'iamt')})
    rows.append({'row': 68, 'label': "Inward supplies from ISD (Central tax)", 'extractor': itc_avl_sum('ISD', 'camt')})
    rows.append({'row': 69, 'label': "Inward supplies from ISD (State/UT tax)", 'extractor': itc_avl_sum('ISD', 'samt')})
    rows.append({'row': 70, 'label': "Inward supplies from ISD (Cess)", 'extractor': itc_avl_sum('ISD', 'csamt')})
    rows.append({'row': 72, 'label': "All other ITC (Integrated tax)", 'extractor': itc_avl_sum('OTH', 'iamt')})
    rows.append({'row': 73, 'label': "All other ITC (Central tax)", 'extractor': itc_avl_sum('OTH', 'camt')})
    rows.append({'row': 74, 'label': "All other ITC (State/UT tax)", 'extractor': itc_avl_sum('OTH', 'samt')})
    rows.append({'row': 75, 'label': "All other ITC (Cess)", 'extractor': itc_avl_sum('OTH', 'csamt')})

    # ITC Reversed
    def itc_rev_sum(ty, field):
        def extractor(d):
            rev = d.get('itc_elg', {}).get('itc_rev', [])
            return sum(item.get(field, 0) for item in rev if item.get('ty') == ty)
        return extractor

    rows.append({'row': 78, 'label': "As per rules 38,42 & 43 (Integrated tax)", 'extractor': itc_rev_sum('RUL', 'iamt')})
    rows.append({'row': 79, 'label': "As per rules 38,42 & 43 (Central tax)", 'extractor': itc_rev_sum('RUL', 'camt')})
    rows.append({'row': 80, 'label': "As per rules 38,42 & 43 (State/UT tax)", 'extractor': itc_rev_sum('RUL', 'samt')})
    rows.append({'row': 81, 'label': "As per rules 38,42 & 43 (Cess)", 'extractor': itc_rev_sum('RUL', 'csamt')})
    rows.append({'row': 83, 'label': "Others (Integrated tax)", 'extractor': itc_rev_sum('OTH', 'iamt')})
    rows.append({'row': 84, 'label': "Others (Central tax)", 'extractor': itc_rev_sum('OTH', 'camt')})
    rows.append({'row': 85, 'label': "Others (State/UT tax)", 'extractor': itc_rev_sum('OTH', 'samt')})
    rows.append({'row': 86, 'label': "Others (Cess)", 'extractor': itc_rev_sum('OTH', 'csamt')})

    # Net ITC (calculated)
    rows.append({'row': 88, 'label': "Net ITC (Integrated tax)", 'calculated': True})
    rows.append({'row': 89, 'label': "Net ITC (Central tax)", 'calculated': True})
    rows.append({'row': 90, 'label': "Net ITC (State/UT tax)", 'calculated': True})
    rows.append({'row': 91, 'label': "Net ITC (Cess)", 'calculated': True})

    # Other Details (D)
    def itc_inelg_sum(ty, field):
        def extractor(d):
            inelg = d.get('itc_elg', {}).get('itc_inelg', [])
            return sum(item.get(field, 0) for item in inelg if item.get('ty') == ty)
        return extractor

    rows.append({'row': 94, 'label': "ITC reclaimed (Integrated tax)", 'extractor': lambda d: 0})
    rows.append({'row': 95, 'label': "ITC reclaimed (Central tax)", 'extractor': lambda d: 0})
    rows.append({'row': 96, 'label': "ITC reclaimed (State/UT tax)", 'extractor': lambda d: 0})
    rows.append({'row': 97, 'label': "ITC reclaimed (Cess)", 'extractor': lambda d: 0})
    rows.append({'row': 99, 'label': "Ineligible ITC u/s 16(4) (Integrated tax)",
                 'extractor': lambda d: itc_inelg_sum('RUL', 'iamt')(d) + itc_inelg_sum('OTH', 'iamt')(d)})
    rows.append({'row': 100, 'label': "Ineligible ITC u/s 16(4) (Central tax)",
                 'extractor': lambda d: itc_inelg_sum('RUL', 'camt')(d) + itc_inelg_sum('OTH', 'camt')(d)})
    rows.append({'row': 101, 'label': "Ineligible ITC u/s 16(4) (State/UT tax)",
                 'extractor': lambda d: itc_inelg_sum('RUL', 'samt')(d) + itc_inelg_sum('OTH', 'samt')(d)})
    rows.append({'row': 102, 'label': "Ineligible ITC u/s 16(4) (Cess)",
                 'extractor': lambda d: itc_inelg_sum('RUL', 'csamt')(d) + itc_inelg_sum('OTH', 'csamt')(d)})

    # Section 5 – values not available, set zero
    for r in [106, 107, 109, 110]:
        rows.append({'row': r, 'label': f"Row {r}", 'extractor': lambda d: 0})

    # 5.1 Interest and Late fee
    add_row(113, "Interest (Integrated tax)", "intr_ltfee", "intr_details", "iamt")
    add_row(114, "Interest (Central tax)", "intr_ltfee", "intr_details", "camt")
    add_row(115, "Interest (State/UT tax)", "intr_ltfee", "intr_details", "samt")
    add_row(116, "Interest (Cess)", "intr_ltfee", "intr_details", "csamt")
    add_row(118, "Late fee (Integrated tax)", "intr_ltfee", "ltfee_details", "iamt")
    add_row(119, "Late fee (Central tax)", "intr_ltfee", "ltfee_details", "camt")
    add_row(120, "Late fee (State/UT tax)", "intr_ltfee", "ltfee_details", "samt")
    add_row(121, "Late fee (Cess)", "intr_ltfee", "ltfee_details", "csamt")

    # 6.1 Payment of tax
    def tax_pay_sum(trancd, tax_type):
        def extractor(d):
            tax_pay = d.get('taxpayble', {}).get('returnsDbCdredList', {}).get('tax_pay', [])
            total = 0
            for item in tax_pay:
                if item.get('trancd') == trancd:
                    total += item.get(tax_type, {}).get('tx', 0)
            return total
        return extractor

    rows.append({'row': 128, 'label': "Tax payable (Other than reverse charge) – Integrated tax",
                 'extractor': tax_pay_sum(30002, 'igst')})
    rows.append({'row': 130, 'label': "Tax payable (Other than reverse charge) – Central tax",
                 'extractor': tax_pay_sum(30002, 'cgst')})
    rows.append({'row': 132, 'label': "Tax payable (Other than reverse charge) – State/UT tax",
                 'extractor': tax_pay_sum(30002, 'sgst')})
    rows.append({'row': 134, 'label': "Tax payable (Other than reverse charge) – Cess",
                 'extractor': tax_pay_sum(30002, 'cess')})

    rows.append({'row': 137, 'label': "Tax payable (Reverse charge) – Integrated tax",
                 'extractor': tax_pay_sum(30003, 'igst')})
    rows.append({'row': 139, 'label': "Tax payable (Reverse charge) – Central tax",
                 'extractor': tax_pay_sum(30003, 'cgst')})
    rows.append({'row': 141, 'label': "Tax payable (Reverse charge) – State/UT tax",
                 'extractor': tax_pay_sum(30003, 'sgst')})
    rows.append({'row': 143, 'label': "Tax payable (Reverse charge) – Cess",
                 'extractor': tax_pay_sum(30003, 'cess')})

    # Total Tax Payable (calculated)
    rows.append({'row': 145, 'label': "Total Tax Payable", 'calculated': True})

    # Tax paid in cash
    def cash_paid_sum(tax_type):
        def extractor(d):
            pd_cash = d.get('taxpayble', {}).get('returnsDbCdredList', {}).get('tax_paid', {}).get('pd_by_cash', [])
            return sum(item.get(tax_type, 0) for item in pd_cash)
        return extractor

    rows.append({'row': 147, 'label': "Tax paid in cash (Other than reverse charge) – Integrated tax",
                 'extractor': cash_paid_sum('igst')})
    rows.append({'row': 148, 'label': "Tax paid in cash (Other than reverse charge) – Central tax",
                 'extractor': cash_paid_sum('cgst')})
    rows.append({'row': 149, 'label': "Tax paid in cash (Other than reverse charge) – State/UT tax",
                 'extractor': cash_paid_sum('sgst')})
    rows.append({'row': 150, 'label': "Tax paid in cash (Other than reverse charge) – Cess",
                 'extractor': cash_paid_sum('cess')})

    rows.append({'row': 152, 'label': "Tax paid in cash (Reverse charge) – Integrated tax",
                 'extractor': cash_paid_sum('igst')})
    rows.append({'row': 153, 'label': "Tax paid in cash (Reverse charge) – Central tax",
                 'extractor': cash_paid_sum('cgst')})
    rows.append({'row': 154, 'label': "Tax paid in cash (Reverse charge) – State/UT tax",
                 'extractor': cash_paid_sum('sgst')})
    rows.append({'row': 155, 'label': "Tax paid in cash (Reverse charge) – Cess",
                 'extractor': cash_paid_sum('cess')})

    # Total Tax Paid in Cash (calculated)
    rows.append({'row': 157, 'label': "Total Tax Paid in Cash", 'calculated': True})   # using row 157 to avoid conflict

    # Tax paid through ITC
    def itc_paid_sum(field):
        def extractor(d):
            pd_itc = d.get('taxpayble', {}).get('returnsDbCdredList', {}).get('tax_paid', {}).get('pd_by_itc', [])
            return sum(item.get(field, 0) for item in pd_itc)
        return extractor

    rows.append({'row': 159, 'label': "Integrated tax paid using Integrated tax",
                 'extractor': itc_paid_sum('igst_igst_amt')})
    rows.append({'row': 160, 'label': "Integrated tax paid using Central tax",
                 'extractor': itc_paid_sum('igst_cgst_amt')})
    rows.append({'row': 161, 'label': "Integrated tax paid using State/UT tax",
                 'extractor': itc_paid_sum('igst_sgst_amt')})
    rows.append({'row': 162, 'label': "Central tax paid using Integrated tax",
                 'extractor': itc_paid_sum('cgst_igst_amt')})
    rows.append({'row': 163, 'label': "Central tax paid using Central tax",
                 'extractor': itc_paid_sum('cgst_cgst_amt')})
    rows.append({'row': 164, 'label': "State/UT tax paid using Integrated tax",
                 'extractor': itc_paid_sum('sgst_igst_amt')})
    rows.append({'row': 165, 'label': "State/UT tax paid using State/UT tax",
                 'extractor': itc_paid_sum('sgst_sgst_amt')})
    rows.append({'row': 166, 'label': "Cess paid using Cess",
                 'extractor': itc_paid_sum('cess_cess_amt')})

    # Total Tax Paid through ITC (calculated)
    rows.append({'row': 168, 'label': "Total Tax Paid through ITC", 'calculated': True})

    # Build data matrix
    month_cols = months_list
    data_matrix = {row['row']: {m: 0 for m in month_cols} for row in rows}

    for month, month_data in months_data.items():
        for row in rows:
            if 'extractor' in row:
                try:
                    val = row['extractor'](month_data)
                    if not isinstance(val, (int, float)):
                        val = 0
                except:
                    val = 0
                data_matrix[row['row']][month] = val

    # Calculate net ITC (rows 88-91)
    for month in month_cols:
        data_matrix[88][month] = (data_matrix[56][month] + data_matrix[59][month] +
                                   data_matrix[62][month] + data_matrix[67][month] +
                                   data_matrix[72][month]) - (data_matrix[78][month] + data_matrix[83][month])
        data_matrix[89][month] = (data_matrix[63][month] + data_matrix[68][month] +
                                   data_matrix[73][month]) - (data_matrix[79][month] + data_matrix[84][month])
        data_matrix[90][month] = (data_matrix[64][month] + data_matrix[69][month] +
                                   data_matrix[74][month]) - (data_matrix[80][month] + data_matrix[85][month])
        data_matrix[91][month] = (data_matrix[57][month] + data_matrix[60][month] +
                                   data_matrix[65][month] + data_matrix[70][month] +
                                   data_matrix[75][month]) - (data_matrix[81][month] + data_matrix[86][month])

        # Total Tax Payable (row 145)
        data_matrix[145][month] = (data_matrix[128][month] + data_matrix[130][month] +
                                    data_matrix[132][month] + data_matrix[134][month] +
                                    data_matrix[137][month] + data_matrix[139][month] +
                                    data_matrix[141][month] + data_matrix[143][month])

        # Total Tax Paid in Cash (row 157) = sum of 147-150 + 152-155
        data_matrix[157][month] = (data_matrix[147][month] + data_matrix[148][month] +
                                    data_matrix[149][month] + data_matrix[150][month] +
                                    data_matrix[152][month] + data_matrix[153][month] +
                                    data_matrix[154][month] + data_matrix[155][month])

        # Total Tax Paid through ITC (row 168) = sum of 159-166
        data_matrix[168][month] = sum(data_matrix[r][month] for r in range(159, 167))

    # Prepare preview rows
    preview_rows = []
    for row in sorted(rows, key=lambda x: x['row']):
        total = sum(data_matrix[row['row']][m] for m in month_cols)
        preview_rows.append({
            'row': row['row'],
            'label': row['label'],
            'values': [data_matrix[row['row']][m] for m in month_cols],
            'total': total
        })

    meta = {
        'gstin': gstin,
        'financial_year': financial_year,
        'no_of_months': no_of_months,
        'form': 'GSTR-3B',
        'months': month_cols,
        'legal_name': legal_name
    }

    return {'meta': meta, 'rows': preview_rows}, None if not errors else "; ".join(errors)

def create_excel_file(consolidated_data, token):
    meta = consolidated_data['meta']
    rows_data = consolidated_data['rows']
    month_cols = meta['months']

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Meta_Data"

    # Meta_Data sheet
    meta_rows = [
        ["Field", "Value", ""],
        ["GSTIN", meta['gstin'], ""],
        ["Financial Year", meta['financial_year'], "Starting month - ending month"],
        ["Legal name of the supplier", meta['legal_name'], ""],
        ["No of Months", meta['no_of_months'], "<No of months JSON file provided>"],
        ["Form", "GSTR-3B", ""],
        ["Creation Date", datetime.now().strftime("%d/%m/%Y %H:%M:%S"), ""]
    ]
    for r in meta_rows:
        ws1.append(r)

    # GSTR-3B sheet
    ws2 = wb.create_sheet("GSTR-3B")

    header_fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Headers
    headers = ["S No.", "Particulars"] + month_cols + ["Total"]
    ws2.append(headers)
    for col_idx, cell in enumerate(ws2[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row_map = {r['row']: r for r in rows_data}

    # Full template rows (must match row numbers used in rows list)
    template_rows = [
        (4, "3.1 Details of outward supplies and inward supplies liable to reverse charge", True),
        (5, "(a) Outward taxable supplies (other than zero, nil & exempted)", False),
        (6, "(i) Total taxable value", False),
        (7, "(ii) Integrated tax", False),
        (8, "(iii) Central tax", False),
        (9, "(iv) State/UT tax", False),
        (10, "(v) Cess", False),
        (11, "(b) Outward taxable supplies (zero rated)", False),
        (12, "(i) Total taxable value", False),
        (13, "(ii) Integrated tax", False),
        (14, "(v) Cess", False),
        (15, "(c) Other outward supplies (nil rated, exempted)", False),
        (16, "(i) Total taxable value", False),
        (17, "(d) Inward supplies (liable to reverse charge)", False),
        (18, "(i) Total taxable value", False),
        (19, "(ii) Integrated tax", False),
        (20, "(iii) Central tax", False),
        (21, "(iv) State/UT tax", False),
        (22, "(v) Cess", False),
        (23, "(e) Non-GST outward supplies", False),
        (24, "(i) Total taxable value", False),
        (25, "", True),  # spacer
        (26, "3.1.1 Details of supplies notified under section 9(5)", True),
        (27, "(i) ECO pays tax u/s 9(5)", False),
        (28, "Total taxable value", False),
        (29, "Integrated tax", False),
        (30, "Central tax", False),
        (31, "State/UT tax", False),
        (32, "Cess", False),
        (33, "(ii) Supplies by registered person through ECO", False),
        (34, "Total taxable value", False),
        (35, "Integrated tax", False),
        (36, "Central tax", False),
        (37, "State/UT tax", False),
        (38, "Cess", False),
        (39, "", True),
        (40, "3.2 Inter State Supplies", True),
        (41, "Supplies made to Unregistered Persons", False),
        (42, "(i) Total taxable value", False),
        (43, "(ii) Integrated tax", False),
        (44, "Supplies made to Composition Taxable Persons", False),
        (45, "(i) Total taxable value", False),
        (46, "(ii) Integrated tax", False),
        (47, "Supplies made to UIN holders", False),
        (48, "(i) Total taxable value", False),
        (49, "(ii) Integrated tax", False),
        (50, "", True),
        (51, "4. Eligible ITC", True),
        (52, "A. ITC Available (whether in full or part)", False),
        (53, "1 Import of goods", False),
        (54, "(i) Integrated tax", False),
        (55, "(ii) Cess", False),
        (56, "2 Import of services", False),
        (57, "(i) Integrated tax", False),
        (58, "(ii) Cess", False),
        (59, "3 Inward supplies liable to reverse charge", False),
        (60, "(i) Integrated tax", False),
        (61, "(ii) Central tax", False),
        (62, "(iii) State/UT tax", False),
        (63, "(iv) Cess", False),
        (64, "4 Inward supplies from ISD", False),
        (65, "(i) Integrated tax", False),
        (66, "(ii) Central tax", False),
        (67, "(iii) State/UT tax", False),
        (68, "(iv) Cess", False),
        (69, "5 All other ITC", False),
        (70, "(i) Integrated tax", False),
        (71, "(ii) Central tax", False),
        (72, "(iii) State/UT tax", False),
        (73, "(iv) Cess", False),
        (74, "B. ITC Reversed", False),
        (75, "1 As per rules 38,42 & 43 and section 17(5)", False),
        (76, "(i) Integrated tax", False),
        (77, "(ii) Central tax", False),
        (78, "(iii) State/UT tax", False),
        (79, "(iv) Cess", False),
        (80, "2 Others", False),
        (81, "(i) Integrated tax", False),
        (82, "(ii) Central tax", False),
        (83, "(iii) State/UT tax", False),
        (84, "(iv) Cess", False),
        (85, "C. Net ITC available (A-B)", False),
        (86, "(i) Integrated tax", False),
        (87, "(ii) Central tax", False),
        (88, "(iii) State/UT tax", False),
        (89, "(iv) Cess", False),
        (90, "D. Other Details", False),
        (91, "1 ITC reclaimed Table 4(B)(2)", False),
        (92, "(i) Integrated tax", False),
        (93, "(ii) Central tax", False),
        (94, "(iii) State/UT tax", False),
        (95, "(iv) Cess", False),
        (96, "2 Ineligible ITC section 16(4)", False),
        (97, "(i) Integrated tax", False),
        (98, "(ii) Central tax", False),
        (99, "(iii) State/UT tax", False),
        (100, "(iv) Cess", False),
        (101, "", True),
        (102, "5. Values of exempt, nil-rated and non-GST inward supplies", True),
        (103, "1 From a supplier under composition scheme, Exempt, Nil rated supply", False),
        (104, "(i) Inter-State supplies", False),
        (105, "(ii) Intra-State supplies", False),
        (106, "2 Non GST supply", False),
        (107, "(i) Inter-State supplies", False),
        (108, "(ii) Intra-State supplies", False),
        (109, "5.1 Interest and Late fee for previous tax period", True),
        (110, "1 Interest Paid", False),
        (111, "(i) Integrated tax", False),
        (112, "(ii) Central tax", False),
        (113, "(iii) State/UT tax", False),
        (114, "(iv) Cess", False),
        (115, "2 Late Fee", False),
        (116, "(i) Integrated tax", False),
        (117, "(ii) Central tax", False),
        (118, "(iii) State/UT tax", False),
        (119, "(iv) Cess", False),
        (120, "", True),
        (121, "6.1 Payment of tax", True),
        (122, "6.1 Tax payable", False),
        (123, "6.1(A) Other than reverse charge", False),
        (124, "1 Integrated tax", False),
        (125, "Tax payable", False),
        (126, "2 Central tax", False),
        (127, "Tax payable", False),
        (128, "3 State/UT tax", False),
        (129, "Tax payable", False),
        (130, "4 Cess", False),
        (131, "Tax payable", False),
        (132, "6.1(B) Reverse charge and supplies made u/s 9(5)", False),
        (133, "1 Integrated tax", False),
        (134, "Tax payable", False),
        (135, "2 Central tax", False),
        (136, "Tax payable", False),
        (137, "3 State/UT tax", False),
        (138, "Tax payable", False),
        (139, "4 Cess", False),
        (140, "Tax payable", False),
        (141, "Total Tax Payable", False),          # data row
        (142, "", True),
        (143, "6.1(B)(ii) Tax paid in Cash (Other than reverse charge)", False),
        (144, "(i) Integrated tax Paid", False),
        (145, "(ii) Central tax Paid", False),
        (146, "(iii) State/UT tax Paid", False),
        (147, "(iv) Cess Paid", False),
        (148, "6.1(B)(ii) Tax paid in Cash (Reverse charge and supplies made u/s 9(5))", False),
        (149, "(i) Integrated tax Paid", False),
        (150, "(ii) Central tax Paid", False),
        (151, "(iii) State/UT tax Paid", False),
        (152, "(iv) Cess Paid", False),
        (153, "Total Tax Paid in Cash", False),      # data row (we'll use row 157 in data, but display here)
        (154, "", True),
        (155, "6.1(A)(i) Tax paid through ITC", False),
        (156, "(i) Integrated tax paid using Integrated tax", False),
        (157, "(ii) Integrated tax paid using Central tax", False),
        (158, "(iii) Integrated tax paid using State/UT tax", False),
        (159, "(iv) Central tax paid using Integrated tax", False),
        (160, "(v) Central tax paid using Central tax", False),
        (161, "(vi) State/UT tax paid using Integrated tax", False),
        (162, "(vii) State/UT tax paid using State/UT tax", False),
        (163, "(viii) Cess Paid using Cess", False),
        (164, "Total Tax Paid through ITC", False),  # data row (we'll use row 168 in data)
    ]

    for row_num, label, is_header in template_rows:
        if is_header:
            ws2.append([label] + [''] * (len(month_cols) + 1))
            # Merge cells for header row
            if len(month_cols) + 2 > 1:
                ws2.merge_cells(start_row=ws2.max_row, start_column=1,
                                end_row=ws2.max_row, end_column=len(month_cols)+3)
            cell = ws2.cell(row=ws2.max_row, column=1)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.font = Font(bold=True, color="1E3C72")
            cell.border = border
        else:
            # Data row
            s_no = row_num - 3  # approximate S.No.
            row_data = [s_no, label]
            for m in month_cols:
                # Look up data by the row number we have in row_map (may be different for totals)
                # For rows that are totals but have different internal row numbers, we map them.
                if row_num == 141:
                    val_row = 145   # Total Tax Payable
                elif row_num == 153:
                    val_row = 157   # Total Tax Paid in Cash
                elif row_num == 164:
                    val_row = 168   # Total Tax Paid through ITC
                else:
                    val_row = row_num
                if val_row in row_map:
                    val = row_map[val_row]['values'][month_cols.index(m)]
                else:
                    val = 0
                row_data.append(val)
            # Total column
            if row_num == 141:
                total_row = 145
            elif row_num == 153:
                total_row = 157
            elif row_num == 164:
                total_row = 168
            else:
                total_row = row_num
            if total_row in row_map:
                total = row_map[total_row]['total']
            else:
                total = 0
            row_data.append(total)
            ws2.append(row_data)

    # Apply borders to all cells
    for row in ws2.iter_rows():
        for cell in row:
            cell.border = border

    timestamp = datetime.now().strftime("%d%m%Y_%H%M%S")
    fname = f"consolidated_GSTR-3B_{timestamp}.xlsx"
    out_path = os.path.join(app.config['OUTPUT_FOLDER'], fname)
    wb.save(out_path)
    return fname

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    cleanup_old_files()
    if 'files[]' not in request.files:
        return jsonify({'error': 'No files uploaded'}), 400

    files = request.files.getlist('files[]')
    saved_paths = []
    for file in files:
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(path)
            saved_paths.append(path)

    if not saved_paths:
        return jsonify({'error': 'No valid JSON files'}), 400

    consolidated, err = parse_gstr3b_files(saved_paths)
    if err:
        return jsonify({'error': err}), 400

    token = str(uuid.uuid4())
    out_fname = create_excel_file(consolidated, token)
    if not hasattr(app, 'file_map'):
        app.file_map = {}
    app.file_map[token] = out_fname

    return jsonify({
        'success': True,
        'preview': consolidated,
        'token': token,
        'fileCount': len(saved_paths),
        'sheetCount': 2,
        'rowCount': len(consolidated['rows']),
        'columnCount': len(consolidated['meta']['months']) + 2
    })

@app.route('/download')
def download():
    token = request.args.get('token')
    if not token or not hasattr(app, 'file_map') or token not in app.file_map:
        return jsonify({'error': 'Invalid token'}), 400
    fname = app.file_map[token]
    path = os.path.join(app.config['OUTPUT_FOLDER'], fname)
    if not os.path.exists(path):
        return jsonify({'error': 'File not found'}), 404
    return send_file(path, as_attachment=True, download_name=fname)

if __name__ == '__main__':
    app.run(host='127.199.99.20', port=5000, debug=True)
