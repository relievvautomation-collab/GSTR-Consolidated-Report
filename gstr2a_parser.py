import os
import json
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import utils

# ---------- Helper functions ----------
def fp_to_month_str(fp):
    """Convert '012025' to 'Jan-25' using utils.format_month."""
    return utils.format_month(fp)

def safe_float(val):
    try:
        return float(val) if val not in (None, '') else 0.0
    except:
        return 0.0

# ---------- Main parsing function ----------
def parse_gstr2a_files(file_paths):
    """
    Returns (consolidated_dict, error_message)
    consolidated_dict = {
        'meta': {'gstin': str, 'financial_year': str, 'months': list},
        'sheets': {
            'b2b': {'columns': list, 'rows': [list of values]},
            'cdn': {'columns': list, 'rows': [list of values]},
            'tcs': {'columns': list, 'rows': [list of values]}
        }
    }
    """
    consolidated = {
        'meta': {'gstin': None, 'financial_year': '', 'months': []},
        'sheets': {
            'b2b': {'columns': [], 'rows': []},
            'cdn': {'columns': [], 'rows': []},
            'tcs': {'columns': [], 'rows': []}
        }
    }

    # Define column headers exactly as in sample output
    b2b_cols = [
        'Month', 'GSTIN of supplier', 'Trade/Legal name', 'Invoice number',
        'Invoice Type', 'Invoice date', 'Invoice Value (₹)', 'Place of supply',
        'Reverse Charge', 'Rate (%)', 'Taxable Value (₹)', 'Integrated Tax (₹)',
        'Central Tax (₹)', 'State/UT Tax (₹)', 'Cess (₹)',
        'GSTR-1/IFF/GSTR-1A/5 Filing Status', 'GSTR-1/IFF/GSTR-1A/5 Filing Date',
        'GSTR-1/IFF/GSTR-1A/5 Filing Period', 'GSTR-3B Filing Status',
        'Amendment made, if any', 'Effective date of cancellation',
        'Source', 'IRN', 'IRN Date'
    ]
    cdn_cols = [
        'Month', 'GSTIN of Supplier', 'Trade/Legal name of the supplier',
        'Note type', 'Note number', 'Note Supply type', 'Note date',
        'Note Value (₹)', 'Place of supply', 'Supply Attract Reverse Charge',
        'Rate (%)', 'Taxable Value (₹)', 'Integrated Tax (₹)',
        'Central Tax (₹)', 'State Tax (₹)', 'Cess Amount (₹)',
        'GSTR-1/IFF/GSTR-1A/5 Filing Status', 'GSTR-1/IFF/GSTR-1A/5 Filing Date',
        'GSTR-1/IFF/GSTR-1A/5 Filing Period', 'GSTR-3B Filing Status',
        'Amendment made, if any', 'Source', 'IRN', 'IRN date'
    ]
    tcs_cols = [
        'Month', 'Tax period GSTR-8', 'Value of supplies returned (₹)',
        'Integrated Tax (₹)', 'Central Tax (₹)', 'State/UT Tax (₹)'
    ]

    consolidated['sheets']['b2b']['columns'] = b2b_cols
    consolidated['sheets']['cdn']['columns'] = cdn_cols
    consolidated['sheets']['tcs']['columns'] = tcs_cols

    months_set = set()
    gstin_first = None

    for file_path in file_paths:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # GSTIN consistency check
        curr_gstin = data.get('gstin')
        if not curr_gstin:
            return None, f"Missing 'gstin' in file {os.path.basename(file_path)}"
        if gstin_first is None:
            gstin_first = curr_gstin
        elif curr_gstin != gstin_first:
            return None, f"GSTIN mismatch: {curr_gstin} != {gstin_first}"

        # Month from 'fp'
        fp = data.get('fp', '')
        month_str = fp_to_month_str(fp) if fp else ''
        months_set.add(month_str)

        # ---------- B2B ----------
        for supplier in data.get('b2b', []):
            ctin = supplier.get('ctin', '')
            cfs = supplier.get('cfs', '')
            cfs3b = supplier.get('cfs3b', '')
            fldtr1 = supplier.get('fldtr1', '')
            flprdr1 = supplier.get('flprdr1', '')

            for inv in supplier.get('inv', []):
                # Aggregate item details
                txval_total = 0.0
                iamt_total = 0.0
                camt_total = 0.0
                samt_total = 0.0
                csamt_total = 0.0
                rate = 0
                for itm in inv.get('itms', []):
                    det = itm.get('itm_det', {})
                    txval_total += safe_float(det.get('txval'))
                    iamt_total += safe_float(det.get('iamt'))
                    camt_total += safe_float(det.get('camt'))
                    samt_total += safe_float(det.get('samt'))
                    csamt_total += safe_float(det.get('csamt'))
                    if rate == 0 and det.get('rt'):
                        rate = det.get('rt')

                row = [
                    month_str,                                 # Month
                    ctin,                                      # GSTIN of supplier
                    '',                                        # Trade/Legal name (N/A)
                    inv.get('inum', ''),                       # Invoice number
                    inv.get('inv_typ', ''),                    # Invoice Type
                    inv.get('idt', ''),                         # Invoice date
                    safe_float(inv.get('val')),                 # Invoice Value
                    inv.get('pos', ''),                          # Place of supply
                    inv.get('rchrg', 'N'),                       # Reverse Charge
                    rate,                                        # Rate (%)
                    txval_total,                                 # Taxable Value
                    iamt_total,                                  # Integrated Tax
                    camt_total,                                  # Central Tax
                    samt_total,                                  # State/UT Tax
                    csamt_total,                                 # Cess
                    cfs,                                         # GSTR-1 Filing Status
                    fldtr1,                                      # GSTR-1 Filing Date
                    flprdr1,                                     # GSTR-1 Filing Period
                    cfs3b,                                       # GSTR-3B Filing Status
                    '',                                          # Amendment made
                    '',                                          # Effective date of cancellation
                    'E-Invoice' if inv.get('irn') else '',      # Source
                    inv.get('irn', ''),                          # IRN
                    inv.get('irngendate', '')                    # IRN Date
                ]
                # Map place of supply
                row[7] = utils.map_place_of_supply(row[7])
                consolidated['sheets']['b2b']['rows'].append(row)

        # ---------- CDN ----------
        for cdn_entry in data.get('cdn', []):
            ctin = cdn_entry.get('ctin', '')
            cfs = cdn_entry.get('cfs', '')
            cfs3b = cdn_entry.get('cfs3b', '')
            fldtr1 = cdn_entry.get('fldtr1', '')
            flprdr1 = cdn_entry.get('flprdr1', '')

            for nt in cdn_entry.get('nt', []):
                # Aggregate item details for note
                txval_total = 0.0
                iamt_total = 0.0
                camt_total = 0.0
                samt_total = 0.0
                csamt_total = 0.0
                rate = 0
                for itm in nt.get('itms', []):
                    det = itm.get('itm_det', {})
                    txval_total += safe_float(det.get('txval'))
                    iamt_total += safe_float(det.get('iamt'))
                    camt_total += safe_float(det.get('camt'))
                    samt_total += safe_float(det.get('samt'))
                    csamt_total += safe_float(det.get('csamt'))
                    if rate == 0 and det.get('rt'):
                        rate = det.get('rt')

                row = [
                    month_str,                                 # Month
                    ctin,                                      # GSTIN of Supplier
                    '',                                        # Trade/Legal name
                    nt.get('ntty', ''),                         # Note type
                    nt.get('nt_num', ''),                       # Note number
                    nt.get('inv_typ', ''),                      # Note Supply type
                    nt.get('nt_dt', ''),                        # Note date
                    safe_float(nt.get('val')),                  # Note Value
                    nt.get('pos', ''),                          # Place of supply
                    nt.get('rchrg', 'N'),                        # Supply Attract Reverse Charge
                    rate,                                        # Rate (%)
                    txval_total,                                 # Taxable Value
                    iamt_total,                                  # Integrated Tax
                    camt_total,                                  # Central Tax
                    samt_total,                                  # State Tax
                    csamt_total,                                 # Cess
                    cfs,                                         # GSTR-1 Filing Status
                    fldtr1,                                      # GSTR-1 Filing Date
                    flprdr1,                                     # GSTR-1 Filing Period
                    cfs3b,                                       # GSTR-3B Filing Status
                    '',                                          # Amendment made
                    'E-Invoice' if nt.get('irn') else '',      # Source
                    nt.get('irn', ''),                           # IRN
                    nt.get('irngendate', '')                     # IRN date
                ]
                # Map place of supply
                row[8] = utils.map_place_of_supply(row[8])
                consolidated['sheets']['cdn']['rows'].append(row)

        # ---------- TCS ----------
        for tcs_entry in data.get('tcs', []):
            row = [
                month_str,                                      # Month
                '',                                             # Tax period GSTR-8 (not available)
                'Not Found',                                    # Value of supplies returned (placeholder)
                safe_float(tcs_entry.get('iamt')),              # Integrated Tax
                safe_float(tcs_entry.get('camt')),              # Central Tax
                safe_float(tcs_entry.get('samt'))               # State/UT Tax
            ]
            consolidated['sheets']['tcs']['rows'].append(row)

    # Meta data
    consolidated['meta']['gstin'] = gstin_first
    months_list = sorted([m for m in months_set if m])
    consolidated['meta']['months'] = months_list
    if months_list:
        # Derive financial year from first month (original fp)
        # For simplicity, we'll store the formatted months; financial year remains unchanged.
        fp_first = datetime.strptime(months_list[0], '%b-%y').strftime('%m%Y')
        consolidated['meta']['financial_year'] = utils.month_to_fy(fp_first)  # need month_to_fy in utils? We'll add it.

    return consolidated, None

# ---------- Excel generation ----------
def create_gstr2a_excel_file(consolidated):
    wb = Workbook()

    # Meta_Data sheet
    ws_meta = wb.active
    ws_meta.title = 'Meta_Data'
    ws_meta.append(['Field', 'Value'])
    ws_meta.append(['GSTIN', consolidated['meta']['gstin']])
    ws_meta.append(['Financial Year', consolidated['meta']['financial_year']])
    ws_meta.append(['No of Months', len(consolidated['meta']['months'])])
    ws_meta.append(['Form', 'GSTR-2A'])
    ws_meta.append(['Creation Date', datetime.now().strftime('%d/%m/%Y %H:%M:%S')])  # includes time
    utils.style_header(ws_meta)
    utils.auto_width(ws_meta)

    # B2B sheet
    ws_b2b = wb.create_sheet('B2B')
    ws_b2b.append(consolidated['sheets']['b2b']['columns'])
    for row in consolidated['sheets']['b2b']['rows']:
        ws_b2b.append(row)
    utils.style_header(ws_b2b)
    utils.apply_number_format(ws_b2b)
    utils.auto_width(ws_b2b)
    utils.freeze_first_row(ws_b2b)

    # CDN sheet
    ws_cdn = wb.create_sheet('CDN')
    ws_cdn.append(consolidated['sheets']['cdn']['columns'])
    for row in consolidated['sheets']['cdn']['rows']:
        ws_cdn.append(row)
    utils.style_header(ws_cdn)
    utils.apply_number_format(ws_cdn)
    utils.auto_width(ws_cdn)
    utils.freeze_first_row(ws_cdn)

    # TCS sheet
    ws_tcs = wb.create_sheet('TCS')
    ws_tcs.append(consolidated['sheets']['tcs']['columns'])
    for row in consolidated['sheets']['tcs']['rows']:
        ws_tcs.append(row)
    utils.style_header(ws_tcs)
    utils.apply_number_format(ws_tcs)
    utils.auto_width(ws_tcs)
    utils.freeze_first_row(ws_tcs)

    # Save file
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    fname = f"GSTR2A_Consolidated_{timestamp}.xlsx"
    out_path = os.path.join('output', fname)
    wb.save(out_path)
    return fname
