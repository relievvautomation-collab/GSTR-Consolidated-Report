import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import utils   # new import

# ------------------------------------------------------------
# Parsing logic
# ------------------------------------------------------------
def parse_gstr1_files(file_paths):
    """
    Parses multiple GSTR-1 JSON files and consolidates data into four sheets:
    b2b (including sez/de), b2cs, cdnr, hsn.
    Returns (consolidated_dict, error_message)
    """
    months_data = {}  # key: period (e.g. '012025'), value: json data
    gstin = None
    errors = []

    for fpath in file_paths:
        try:
            with open(fpath, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            errors.append(f"Error reading {fpath}: {str(e)}")
            continue

        fp = data.get('fp')
        if not fp:
            errors.append(f"Missing 'fp' in {fpath}")
            continue

        months_data[fp] = data
        if gstin is None:
            gstin = data.get('gstin', '')
        elif data.get('gstin') != gstin:
            errors.append(f"GSTIN mismatch in {fpath}")

    if not months_data:
        return None, "No valid JSON files. " + "; ".join(errors)

    # Sort periods and derive financial year
    periods = sorted(months_data.keys())
    first = periods[0]
    last = periods[-1]
    fy_start = first[:2] + first[2:4]
    fy_end = last[:2] + last[2:4]
    financial_year = f"{fy_start} - {fy_end}"

    # Prepare sheet structures
    sheets = {
        'b2b': {'columns': [], 'rows': []},
        'b2cs': {'columns': [], 'rows': []},
        'cdnr': {'columns': [], 'rows': []},
        'hsn': {'columns': [], 'rows': []}
    }

    # Define columns for each sheet (must match sample output)
    sheets['b2b']['columns'] = [
        "Month", "GSTIN/UIN of Recipient", "Receiver Name", "Invoice Number",
        "Invoice date", "Invoice Value", "Place Of Supply", "Reverse Charge",
        "Applicable % of Tax Rate", "Invoice Type", "E-Commerce GSTIN",
        "Rate", "Taxable Value", "IGST Amount", "CGST Amount", "SGST Amount", "Cess Amount"
    ]
    sheets['b2cs']['columns'] = [
        "Month", "Type", "Place Of Supply", "Applicable % of Tax Rate",
        "Rate", "Taxable Value", "IGST Amount", "CGST Amount", "SGST Amount", "Cess Amount"
    ]
    sheets['cdnr']['columns'] = [
        "Month", "GSTIN/UIN of Recipient", "Receiver Name", "Note Number",
        "Note Date", "Note Type", "Place Of Supply", "Reverse Charge",
        "Note Supply Type", "Note Value", "Applicable % of Tax Rate",
        "Rate", "Taxable Value", "IGST Amount", "CGST Amount", "SGST Amount", "Cess Amount"
    ]
    sheets['hsn']['columns'] = [
        "Month", "HSN", "Description", "UQC", "Total Quantity", "Rate",
        "Taxable Value", "Integrated Tax Amount", "Central Tax Amount", "State/UT Tax Amount"
    ]

    # Helper to convert fp to formatted month string "MMM-YY"
    def fp_to_month_str(fp):
        return utils.format_month(fp)

    # Process each month's data
    for period, data in months_data.items():
        month_str = fp_to_month_str(period)

        # ---- B2B (including SEZ/DE) ----
        b2b_list = data.get('b2b', [])
        for b2b_entry in b2b_list:
            ctin = b2b_entry.get('ctin', '')
            inv_list = b2b_entry.get('inv', [])
            for inv in inv_list:
                inv_num = inv.get('inum', '')
                inv_dt = inv.get('idt', '')
                inv_val = inv.get('val', 0)
                pos = inv.get('pos', '')
                rchrg = inv.get('rchrg', 'N')
                inv_typ = inv.get('inv_typ', 'R')
                itms = inv.get('itms', [])
                for item in itms:
                    itm_det = item.get('itm_det', {})
                    rate = itm_det.get('rt', 0)
                    txval = itm_det.get('txval', 0)
                    iamt = itm_det.get('iamt', 0)
                    camt = itm_det.get('camt', 0)
                    samt = itm_det.get('samt', 0)
                    csamt = itm_det.get('csamt', 0)
                    row = [
                        month_str, ctin, "", inv_num, inv_dt, inv_val, pos, rchrg,
                        "", inv_typ, "", rate, txval, iamt, camt, samt, csamt
                    ]
                    # Map place of supply to full state name
                    row[6] = utils.map_place_of_supply(row[6])
                    sheets['b2b']['rows'].append(row)

        # ---- B2CS ----
        b2cs_list = data.get('b2cs', [])
        for b2cs in b2cs_list:
            sply_ty = b2cs.get('sply_ty', '')  # INTER/INTRA
            pos = b2cs.get('pos', '')
            typ = b2cs.get('typ', '')  # OE, etc.
            rate = b2cs.get('rt', 0)
            txval = b2cs.get('txval', 0)
            iamt = b2cs.get('iamt', 0)
            camt = b2cs.get('camt', 0)
            samt = b2cs.get('samt', 0)
            csamt = b2cs.get('csamt', 0)
            row = [
                month_str, sply_ty, pos, typ, rate, txval, iamt, camt, samt, csamt
            ]
            row[2] = utils.map_place_of_supply(row[2])
            sheets['b2cs']['rows'].append(row)

        # ---- CDNR ----
        cdnr_list = data.get('cdnr', [])
        for cdnr_entry in cdnr_list:
            ctin = cdnr_entry.get('ctin', '')
            nt_list = cdnr_entry.get('nt', [])
            for nt in nt_list:
                nt_num = nt.get('nt_num', '')
                nt_dt = nt.get('nt_dt', '')
                ntty = nt.get('ntty', '')  # C/D
                pos = nt.get('pos', '')
                rchrg = nt.get('rchrg', 'N')
                inv_typ = nt.get('inv_typ', 'R')
                val = nt.get('val', 0)
                itms = nt.get('itms', [])
                for item in itms:
                    itm_det = item.get('itm_det', {})
                    rate = itm_det.get('rt', 0)
                    txval = itm_det.get('txval', 0)
                    iamt = itm_det.get('iamt', 0)
                    camt = itm_det.get('camt', 0)
                    samt = itm_det.get('samt', 0)
                    csamt = itm_det.get('csamt', 0)
                    row = [
                        month_str, ctin, "", nt_num, nt_dt, ntty, pos, rchrg,
                        inv_typ, val, "", rate, txval, iamt, camt, samt, csamt
                    ]
                    row[6] = utils.map_place_of_supply(row[6])
                    sheets['cdnr']['rows'].append(row)

        # ---- HSN ----
        hsn_obj = data.get('hsn', {})
        hsn_data = hsn_obj.get('data', [])
        for hsn_entry in hsn_data:
            hsn_sc = hsn_entry.get('hsn_sc', '')
            desc = hsn_entry.get('desc', '')
            uqc = hsn_entry.get('uqc', '')
            qty = hsn_entry.get('qty', 0)
            rate = hsn_entry.get('rt', 0)
            txval = hsn_entry.get('txval', 0)
            iamt = hsn_entry.get('iamt', 0)
            camt = hsn_entry.get('camt', 0)
            samt = hsn_entry.get('samt', 0)
            row = [
                month_str, hsn_sc, desc, uqc, qty, rate, txval, iamt, camt, samt
            ]
            sheets['hsn']['rows'].append(row)

    meta = {
        'gstin': gstin,
        'financial_year': financial_year,
        'no_of_months': len(periods),
        'form': 'GSTR-1',
        'months': periods
    }

    return {'meta': meta, 'sheets': sheets}, None if not errors else "; ".join(errors)

# ------------------------------------------------------------
# Excel creation
# ------------------------------------------------------------
def create_gstr1_excel_file(consolidated_data):
    meta = consolidated_data['meta']
    sheets_data = consolidated_data['sheets']

    wb = Workbook()
    # Meta_Data sheet
    ws_meta = wb.active
    ws_meta.title = "Meta_Data"
    meta_rows = [
        ["Field", "Value"],
        ["GSTIN", meta['gstin']],
        ["Financial Year", meta['financial_year']],
        ["No of Months", meta['no_of_months']],
        ["Form", "GSTR-1"],
        ["Creation Date", datetime.now().strftime("%d/%m/%Y %H:%M:%S")]  # includes time
    ]
    for r in meta_rows:
        ws_meta.append(r)
    utils.style_header(ws_meta)
    utils.auto_width(ws_meta)

    # Define border style for data sheets
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Sheet: b2b,sez,de
    if 'b2b' in sheets_data:
        ws_b2b = wb.create_sheet("b2b,sez,de")
        ws_b2b.append(sheets_data['b2b']['columns'])
        # style header
        for col_idx, cell in enumerate(ws_b2b[1], 1):
            cell.fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        # append rows (raw numbers)
        for row in sheets_data['b2b']['rows']:
            ws_b2b.append(row)
        # apply borders and number formatting
        for row in ws_b2b.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
        utils.apply_number_format(ws_b2b)
        utils.auto_width(ws_b2b)
        utils.freeze_first_row(ws_b2b)

    # Sheet: b2cs
    if 'b2cs' in sheets_data:
        ws_b2cs = wb.create_sheet("b2cs")
        ws_b2cs.append(sheets_data['b2cs']['columns'])
        for col_idx, cell in enumerate(ws_b2cs[1], 1):
            cell.fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        for row in sheets_data['b2cs']['rows']:
            ws_b2cs.append(row)
        for row in ws_b2cs.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
        utils.apply_number_format(ws_b2cs)
        utils.auto_width(ws_b2cs)
        utils.freeze_first_row(ws_b2cs)

    # Sheet: cdnr
    if 'cdnr' in sheets_data:
        ws_cdnr = wb.create_sheet("cdnr")
        ws_cdnr.append(sheets_data['cdnr']['columns'])
        for col_idx, cell in enumerate(ws_cdnr[1], 1):
            cell.fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        for row in sheets_data['cdnr']['rows']:
            ws_cdnr.append(row)
        for row in ws_cdnr.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
        utils.apply_number_format(ws_cdnr)
        utils.auto_width(ws_cdnr)
        utils.freeze_first_row(ws_cdnr)

    # Sheet: hsn
    if 'hsn' in sheets_data:
        ws_hsn = wb.create_sheet("hsn")
        ws_hsn.append(sheets_data['hsn']['columns'])
        for col_idx, cell in enumerate(ws_hsn[1], 1):
            cell.fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        for row in sheets_data['hsn']['rows']:
            ws_hsn.append(row)
        for row in ws_hsn.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
        utils.apply_number_format(ws_hsn)
        utils.auto_width(ws_hsn)
        utils.freeze_first_row(ws_hsn)

    timestamp = datetime.now().strftime("%d%m%Y_%H%M%S")
    fname = f"consolidated_GSTR-1_{timestamp}.xlsx"
    out_path = os.path.join('output', fname)
    wb.save(out_path)
    return fname
