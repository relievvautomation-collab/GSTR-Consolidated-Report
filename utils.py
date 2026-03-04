import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill, Font

# Global cache for POS mapping
_pos_map = None

def get_pos_master():
    """
    Load POS Master.xlsx and return a dictionary mapping two-digit state code
    to the full string (e.g., '07' → '07 - Delhi').
    """
    global _pos_map
    if _pos_map is None:
        _pos_map = {}
        try:
            wb = openpyxl.load_workbook('POS Master.xlsx', data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = row[0]      # column A (numeric)
                name = row[1]      # column B (full string)
                if code is not None and name is not None:
                    try:
                        code_int = int(float(code))
                        code_str = f"{code_int:02d}"
                        _pos_map[code_str] = name
                    except (ValueError, TypeError):
                        pass
        except Exception as e:
            print(f"Warning: Could not load POS Master: {e}")
    return _pos_map

def map_place_of_supply(pos_code):
    """
    Convert a place‑of‑supply value like '07' or '07N' to the full state name
    using the POS master. If not found, return the original value.
    """
    if not pos_code:
        return pos_code
    # Take first two digits
    code = str(pos_code).strip()[:2]
    mapping = get_pos_master()
    return mapping.get(code, pos_code)

def format_month(fp):
    """
    Convert a period string like '012025' or '2025-01-01 00:00:00'
    to 'MMM-YY' (e.g., 'Jan-25').
    """
    if not fp:
        return ''
    try:
        # If it's a six-digit period (MMYYYY)
        if len(fp) == 6 and fp.isdigit():
            dt = datetime.strptime(fp, '%m%Y')
        # If it's the old datetime string (YYYY-MM-DD ...)
        elif ' ' in fp:
            dt = datetime.strptime(fp.split()[0], '%Y-%m-%d')
        else:
            return fp
        return dt.strftime('%b-%y')   # e.g., Jan-25
    except:
        return fp

def style_header(ws, row_idx=1):
    """Apply dark blue background and white bold font to the header row."""
    header_fill = PatternFill(start_color='1E3C72', end_color='1E3C72', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[row_idx]:
        cell.fill = header_fill
        cell.font = header_font

def apply_number_format(ws):
    """
    Apply Indian number format (#,##,##0.00) to every numeric cell in the worksheet.
    """
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##,##0.00'

def auto_width(ws, max_width=50):
    """Adjust column widths based on content (capped at max_width)."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

def freeze_first_row(ws):
    """Freeze the first row so it stays visible while scrolling."""
    ws.freeze_panes = 'A2'
    
def month_to_fy(month_str):
    """Convert month like '012025' to financial year e.g. '2024-2025'"""
    try:
        month = int(month_str[0:2])
        year = int(month_str[2:6])
        if month >= 4:
            return f"{year}-{year+1}"
        else:
            return f"{year-1}-{year}"
    except:
        return ''
